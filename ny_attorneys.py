import asyncio
import profile
import re
import time
from asyncio import streams
from email import header
from email.mime import base
from pprint import pprint

import grequests
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, workbook

session = requests.Session()

base_url = 'https://www.martindale.com/'
ny_link = 'https://www.martindale.com/by-location/new-york-lawyers/new-york/'
ny_lawyers_link = 'https://www.martindale.com/all-lawyers/new-york/new-york/?pageSize=500'

referer = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.67 Safari/537.36"
    }

phone_number_re = r'^\(?([0-9]{3})\)?[-.]?([0-9]{3})[-.]?([0-9]{4})$'

def get_resp(link, referer):
    resp = requests.get(link, verify=False, headers=referer)
    if not resp or resp.status_code != 200:
        print("Unable-to-connect-to-{}".format(link))
        return ''
    return resp

def format_contact_details(phone_number):
    formatted_phone_number = ''
    ph_no = re.search(phone_number_re, phone_number, re.I)
    if not ph_no:
        return ''
    formatted_phone_number = "({}) {}-{}".format(ph_no.group(1), ph_no.group(2), ph_no.group(3))
    
    return formatted_phone_number


def write_attorney_details_to_excel(attorney_details):
    wb = load_workbook('Web Scrapping_Sample Records.xlsx')
    sheets = wb.sheetnames
    
    ny_attorney_sheet = wb[sheets[0]]
    row = 6
    for att_detail in attorney_details:
        ny_attorney_sheet.cell(row=row, column=1).value = att_detail['attorney_name'] if 'attorney_name' in att_detail and att_detail['attorney_name'] else ''

        ny_attorney_sheet.cell(row=row, column=2).value = att_detail['attorney_isnl'] if 'attorney_isnl' in att_detail and att_detail['attorney_isnl'] else ''
        
        ny_attorney_sheet.cell(row=row, column=3).value = att_detail['attorney_title'] if 'attorney_title' in att_detail and att_detail['attorney_title'] else ''
        
        ny_attorney_sheet.cell(row=row, column=4).value = att_detail['attorney_company']if 'attorney_company' in att_detail and att_detail['attorney_company'] else ''

        ny_attorney_sheet.cell(row=row, column=5).value = att_detail['attorney_address'] if 'attorney_address' in att_detail and att_detail['attorney_address'] else ''

        ny_attorney_sheet.cell(row=row, column=6).value = att_detail['attorney_phone'] if 'attorney_phone' in att_detail and att_detail['attorney_phone'] else ''

        ny_attorney_sheet.cell(row=row, column=7).value = att_detail['attorney_fax'] if 'attorney_fax' in att_detail and att_detail['attorney_fax'] else '' 

        ny_attorney_sheet.cell(row=row, column=8).value = att_detail['attorney_law_school'] if 'attorney_law_school' in att_detail and att_detail['attorney_law_school'] else ''

        ny_attorney_sheet.cell(row=row, column=9).value = att_detail['attorney_link']
        row += 1
    wb.save('Web Scrapping_Sample Records.xlsx')
    print("Saved-attorney-results-to-excel")
    return True
        
def parse_attorney_details(profile_page_soup, profile_link):
    attorney_detail = {}
    # Attorney Name
    name = profile_page_soup.find('h1', {'class': 'profile-title--bold'})
    if name and name.text:
        attorney_detail['attorney_name'] = name.text.strip()

    # Attorney Title
    title = profile_page_soup.find('li', {'class': 'masthead-list__item masthead-list__item--bold'})
    if title and title.text:
        attorney_detail['attorney_title'] = title.text.strip()

        # Attorney Company
        company = title.find('span')
        if company and company.text:
            attorney_detail['attorney_company'] = company.text.strip()

    # Attorney address
    address = profile_page_soup.find('address')
    if address and address.text:
        attorney_detail['attorney_address'] = address.text.strip()

    # contact details
    contact_details = profile_page_soup.find('div', {'id': 'education-section'})
    if contact_details:
        contact_info_divs = contact_details.findAll('div', {'class': 'row collapse experience-section clearfix'})

        # Contact information
        contact_info_div_spans = contact_info_divs[0].findAll('span')
        if contact_info_div_spans[0] and contact_info_div_spans[0].text:
            phone = contact_info_div_spans[0].text.strip()
            attorney_detail['attorney_phone'] = format_contact_details(phone)
        if len(contact_info_div_spans) > 2 and contact_info_div_spans[2].text:
            fax = contact_info_div_spans[2].text.strip()
            attorney_detail['attorney_fax'] = format_contact_details(fax)

        # University
        if len(contact_info_divs) > 2:
            law_school_details_div = contact_info_divs[2]
            law_school_detail = law_school_details_div.find('div', {'class': 'small-12 medium-9 columns experience-value'})
            if law_school_detail and law_school_detail.text:
                attorney_detail['attorney_law_school'] = law_school_detail.text.strip()

        # ISLN
        isnl_details_div = contact_info_divs.pop()
        isnl_detail = isnl_details_div.find('div', {'class': 'small-12 medium-9 columns experience-value'})
        if isnl_detail and isnl_detail.text:
            attorney_detail['attorney_isnl'] = isnl_detail.text.strip() 

    # profile link
    attorney_detail['attorney_link'] = profile_link

    return attorney_detail

def parse_ny_attorneys():
    ny_attorney_page_resp = session.get(ny_lawyers_link, headers=referer, verify=False)
    if not ny_attorney_page_resp:
        print("Unable-to-reach-ny-lawyers-link-{}".format(ny_lawyers_link))
        return

    ny_attorney_page_soup = BeautifulSoup(ny_attorney_page_resp.content, 'lxml')

    if not ny_attorney_page_soup:
        print("Error-generating-page-html")
    
    attorney_details = []
    i = 0 
    for link in ny_attorney_page_soup.findAll('a', {'class': 'opt-d-title'}):
        profile_link = ''
        profile_link = link['href']
        profile_resp = session.get(profile_link, verify=False, headers=referer)
        print("PROFILEEEEEE ====> {} ---- {}".format(i, profile_resp.status_code))
        i +=1
        profile_page_soup = BeautifulSoup(profile_resp.content, 'lxml')

        attorney_detail = parse_attorney_details(profile_page_soup, profile_link)
        if attorney_detail:
            attorney_details.append(attorney_detail)
        time.sleep(5)
   
    write_attorney_details_to_excel(attorney_details)

if __name__ == '__main__':
    parse_ny_attorneys()
